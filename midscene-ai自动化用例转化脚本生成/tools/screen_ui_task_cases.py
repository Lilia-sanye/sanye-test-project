# -*- coding: utf-8 -*-
"""Screen xlsx/csv for Midscene功能用例筛选; write midscene自动化输出/{cases,report,drafts}."""
from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]


def resolve_midscene_skill_root(project_root: Path) -> Path:
    """优先 .cursor/skills/Midscene，其次 .trae/skills/Midscene。"""
    for sub in (".cursor", ".trae"):
        p = project_root / sub / "skills" / "Midscene"
        if p.is_dir():
            return p
    return project_root / ".cursor" / "skills" / "Midscene"


MIDSCENE_ROOT = resolve_midscene_skill_root(ROOT)
SKILL_MD = MIDSCENE_ROOT / "Midscene功能用例筛选" / "SKILL.md"
YAML_SKILL_MD = MIDSCENE_ROOT / "Midscene安卓自动化脚本" / "SKILL.md"
MIDSCENE_INDEX = MIDSCENE_ROOT / "README.md"
DEFAULT_INPUT = ROOT / "修订版本" / "界面和任务设置.xlsx"
MIDSCENE_OUT = ROOT / "midscene自动化输出"
DEFAULT_SCREEN_OUT = MIDSCENE_OUT
FINAL_YAML_DIR = MIDSCENE_OUT / "scripts"
DRAFTS_DIR = MIDSCENE_OUT / "drafts"

HIGH = [
    "点击", "tap", "touch", "输入", "type", "input", "选择", "select", "choose",
    "滑动", "swipe", "scroll", "等待", "wait", "sleep",
]
MID = ["验证", "verify", "assert", "检查", "check", "validate", "显示", "show", "display"]
UI_EXTRA = [
    "进入", "打开", "关闭", "切换", "勾选", "长按", "双击", "拖动", "拖拽", "缩放",
    "设置", "修改", "确认", "返回", "退出", "登录", "保存", "删除", "新建", "创建",
    "添加", "操作", "执行", "选择时区", "手动选择", "手动修改", "返航", "充电",
]
HARD_BLOCK_PATTERNS = [
    (r"adb\s+shell|ADB命令|通过ADB", "ADB/命令行"),
    (r"登录Linux|Linux从板|终端执行|sudo\s+|timedatectl|/etc/timezone", "Linux/双板终端"),
    (r"金属遮罩|拔卡|遮挡|屏蔽.*信号", "物理/射频环境"),
    (r"重启设备|\breboot\b|开机后.*分钟", "长时重启/等待"),
    (r"重复测试\d+次|记录.*偏差", "人工多次采样"),
]
VISUAL_DIRS = ("状态栏", "主界面", "地图展示", "快捷栏", "UI/UE", "快速任务")
ROBOT_DOMAIN_KW = (
    "满袋", "倒草", "断点", "规划", "回充", "草箱", "位图", "割吸", "收集", "机载", "整机", "实机", "导航",
)

# 逻辑字段 → 表头候选（按优先级）
HEADER_ALIASES: dict[str, list[str]] = {
    "用例ID": ["用例ID", "ID", "编号"],
    "UUID": ["UUID"],
    "用例标题": ["用例标题", "标题"],
    "用例目录": ["用例目录", "目录"],
    "所属模块": ["所属模块", "模块"],
    "等级": ["等级", "优先级", "用例等级"],
    "前置条件": ["前置条件", "precondition"],
    "步骤描述类型": ["步骤描述类型"],
    "步骤描述": ["步骤描述", "操作步骤", "测试步骤"],
    "预期结果": ["预期结果", "expected"],
    "用例类型": ["用例类型"],
}

BOILERPLATE_PRE = re.compile(
    r"【测试环境】.*?【记录要求】.*?(?=【业务焦点】|【本用例前提】|$)", re.DOTALL
)
BOILERPLATE_WRAP = re.compile(
    r"^【(?:执行通则|操作步骤|结果判定|预期现象)】", re.MULTILINE
)


def strip_boilerplate(text: str) -> str:
    if not text:
        return ""
    t = BOILERPLATE_PRE.sub("", text)
    for tag in ("【本用例前提】", "【业务焦点】"):
        if tag in t:
            t = t.split(tag)[-1]
    t = BOILERPLATE_WRAP.sub("", t)
    return t.strip().strip("，").strip(",")


def resolve_headers(raw_headers: list[str | None]) -> dict[str, str]:
    """逻辑名 -> 实际列名"""
    headers = [str(h).strip() for h in raw_headers if h is not None and str(h).strip()]
    mapping: dict[str, str] = {}
    for logical, candidates in HEADER_ALIASES.items():
        for c in candidates:
            if c in headers:
                mapping[logical] = c
                break
    return mapping


def row_to_case(row: dict[str, str], col: dict[str, str], *, row_no: int) -> dict:
    def g(logical: str, default: str = "") -> str:
        key = col.get(logical)
        if not key:
            return default
        return strip_boilerplate(str(row.get(key, default) or ""))

    cid = g("用例ID")
    if not cid:
        cid = f"ROW-{row_no}"
    module = g("所属模块") or g("用例目录")
    return {
        "用例ID": cid,
        "UUID": g("UUID"),
        "用例标题": g("用例标题"),
        "用例目录": g("用例目录") or module,
        "所属模块": module,
        "等级": g("等级"),
        "用例类型": g("用例类型"),
        "前置条件": g("前置条件"),
        "步骤描述类型": g("步骤描述类型"),
        "步骤描述": g("步骤描述"),
        "预期结果": g("预期结果"),
    }


def load_cases_from_csv(path: Path) -> list[dict]:
    cases: list[dict] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError(f"CSV 无表头: {path}")
        col = resolve_headers(list(reader.fieldnames))
        if "步骤描述" not in col and "用例标题" not in col:
            raise ValueError(f"无法映射步骤/标题列，表头: {reader.fieldnames}")
        for i, row in enumerate(reader, start=2):
            case = row_to_case(row, col, row_no=i)
            if not case["用例标题"] and not case["步骤描述"]:
                continue
            cases.append(case)
    return cases


def load_cases_from_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    raw_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = resolve_headers(raw_headers)
    if "步骤描述" not in col and "用例标题" not in col:
        raise ValueError(f"无法映射步骤/标题列，表头: {raw_headers}")
    cases: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row_dict: dict[str, str] = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h:
                v = ws.cell(r, c).value
                row_dict[str(h).strip()] = "" if v is None else str(v)
        if col.get("用例ID") and not row_dict.get(col["用例ID"], "").strip():
            continue
        case = row_to_case(row_dict, col, row_no=r)
        if not case["用例标题"] and not case["步骤描述"]:
            continue
        cases.append(case)
    return cases


def load_cases(path: Path) -> list[dict]:
    path = path.resolve()
    suf = path.suffix.lower()
    if suf == ".csv":
        return load_cases_from_csv(path)
    if suf in (".xlsx", ".xls"):
        return load_cases_from_xlsx(path)
    raise ValueError(f"不支持格式: {suf}，请使用 xlsx 或 csv")


def find_kws(text: str, kws: list[str]) -> list[str]:
    t = text.lower()
    return [k for k in kws if k.lower() in t]


def hard_blockers(text: str) -> list[str]:
    return [name for pat, name in HARD_BLOCK_PATTERNS if re.search(pat, text, re.I)]


def is_robot_non_ui(case: dict) -> bool:
    blob = case["用例标题"] + case["步骤描述"] + case.get("所属模块", "") + case.get("用例目录", "")
    if not any(k in blob for k in ROBOT_DOMAIN_KW):
        return False
    st = case["步骤描述"]
    return not find_kws(st, HIGH) and not find_kws(st, UI_EXTRA[:14])


def skill_classify(case: dict) -> tuple[str, str]:
    st = case["步骤描述"]
    high_s = find_kws(st, HIGH)
    mid_s = find_kws(st, MID)
    extra_s = find_kws(st, UI_EXTRA)
    if high_s:
        return "convert_high", "、".join(high_s)
    if mid_s:
        return "convert_mid", "、".join(mid_s)
    if re.search(r"查看|观察|核对|对比", st) and not extra_s:
        return "reject_observe", "步骤仅查看/观察，无技能列表内操作动词"
    if extra_s:
        return "convert_extra", "、".join(extra_s[:5])
    if not st.strip():
        return "reject_empty", "步骤为空"
    return "reject_other", "步骤无点击/输入/选择/验证等关键词"


def is_visual_recoverable(case: dict) -> bool:
    d = case["用例目录"] + case.get("所属模块", "")
    if not any(v in d for v in VISUAL_DIRS):
        return False
    st = case["步骤描述"]
    return bool(re.search(r"查看|观察|显示|状态栏|图标|电量|信号", st + case["用例标题"]))


def assign_tier(case: dict, skill_label: str, match_reason: str) -> tuple[str, str]:
    if is_robot_non_ui(case):
        return "D", "整机/业务域用例且无明确机载 UI 操作，宜保留为功能测试"

    all_text = case["前置条件"] + case["步骤描述"] + case["用例标题"]
    blocks = hard_blockers(all_text)

    if skill_label.startswith("convert"):
        if blocks:
            return "B", "；".join(blocks)
        if skill_label == "convert_mid" and re.search(r"^[^\n]*查看", case["步骤描述"].strip()):
            return "B", "以观察/显示类步骤为主，需环境注入或截图断言"
        return "A", "机载 UI 可 Midscene 执行"

    if skill_label == "reject_observe" and is_visual_recoverable(case):
        return "C", "建议改为 aiAssert/智能等待（机载视觉类）"

    if skill_label == "reject_other":
        st = case["步骤描述"]
        if find_kws(st, UI_EXTRA) or re.search(r"拖动|缩放|添加|按住", st):
            return "C", "无标准关键词但有手势/添加类操作，可补写 ai 步骤"
        if re.search(r"处于|触发|连接充电|信号", st + case["用例标题"]):
            return "C", "状态/环境类步骤，需台架造数 + 视觉断言"
        if "分支主题" in st or len(st.strip()) < 4:
            return "D", "步骤描述不完整"
        return "C", "建议人工补充操作动词后转化"

    if skill_label == "reject_empty":
        return "D", "无步骤描述"

    return "D", match_reason


def steps_to_flow(case: dict) -> list[str]:
    st = case["步骤描述"].replace("\r\n", "\n")
    lines = []
    for part in re.split(r"\n|(?<=[；;])|(?<=\d[、.])", st):
        part = re.sub(r"^\d+[、.]\s*", "", part.strip())
        if len(part) >= 2:
            lines.append(part)
    if not lines:
        lines = [st.strip()] if st.strip() else [case["用例标题"]]
    flow = [f"ai: {line}" for line in lines[:8]]
    exp = case["预期结果"].split("\n")[0].strip()
    if exp:
        flow.append(f"aiAssert: {exp[:120]}")
    return flow


def detect_shared_steps(cases: list[dict]) -> list[dict]:
    patterns = [
        ("登录", r"登录"),
        ("进入系统设置", r"系统设置|日期和时间"),
        ("返回主界面", r"返回|主界面|主页"),
    ]
    rows = []
    for name, pat in patterns:
        hits = [c for c in cases if re.search(pat, c["步骤描述"] + c["前置条件"])]
        if len(hits) >= 2:
            rows.append(
                {
                    "共享步骤名": name,
                    "出现次数": len(hits),
                    "示例用例ID": hits[0]["用例ID"],
                    "示例标题": hits[0]["用例标题"][:40],
                }
            )
    return rows


def module_key(case: dict) -> str:
    d = case.get("用例目录") or case.get("所属模块", "")
    parts = d.split("/")
    if len(parts) >= 4:
        return "/".join(parts[2:4])
    return d[:48] if d else "未分类"


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def build_yaml_draft(a_cases: list[dict], limit: int = 15, *, stem: str = "筛选套件") -> str:
    lines = [
        "# 草稿：screen_ui_task_cases.py 生成 — 须审阅后移至 midscene自动化输出/scripts/",
        "# 本文件位于 drafts/，不可 midscene run",
        f"# 精修: {YAML_SKILL_MD.relative_to(ROOT).as_posix()}",
        "config:",
        "  ai:",
        "    actionTimeout: 30000",
        "    retryInterval: 200",
        "  page:",
        "    defaultWait: 200",
        "",
        "android:",
        '  deviceId: "<设备ID>"',
        '  androidAdbPath: "D:/soft/android-sdk/platform-tools/adb.exe"',
        "",
        "agent:",
        f'  testId: "{stem}-P0-A档冒烟"',
        f'  groupName: "{stem}"',
        "  generateReport: true",
        '  aiActContext: "权限弹窗点击允许；更新提示跳过；确认框点确认"',
        "",
        "tasks:",
    ]
    for case in a_cases[:limit]:
        name = case["用例标题"][:36].replace('"', "'")
        lines.append(f'  - name: "{case["用例ID"]}-{name}"')
        lines.append("    repeat: 1")
        lines.append("    flow:")
        for step in steps_to_flow(case):
            lines.append(f"      - {step}")
        lines.append("")
    return "\n".join(lines) + "\n"


def main(
    input_path: Path | None = None,
    out_root: Path | None = None,
    stem: str | None = None,
) -> None:
    input_path = (input_path or DEFAULT_INPUT).resolve()
    stem = stem or input_path.stem
    out_root = (out_root or DEFAULT_SCREEN_OUT).resolve()
    report_dir = out_root / "report"
    cases_dir = out_root / "cases"
    drafts_dir = out_root / "drafts"

    cases = load_cases(input_path)
    enriched = []
    tier_counts: Counter[str] = Counter()
    skill_counts: Counter[str] = Counter()

    for case in cases:
        skill_label, match_reason = skill_classify(case)
        tier, tier_note = assign_tier(case, skill_label, match_reason)
        skill_counts[skill_label] += 1
        tier_counts[tier] += 1
        manual_accept = ""
        if tier == "B":
            manual_accept = "；".join(hard_blockers(case["前置条件"] + case["步骤描述"] + case["预期结果"]))
        enriched.append(
            {
                **case,
                "技能分类": skill_label,
                "匹配说明": match_reason,
                "执行档位": tier,
                "档位说明": tier_note,
                "人工验收项": manual_accept,
                "模块": module_key(case),
            }
        )

    auto_rows = [r for r in enriched if r["执行档位"] in ("A", "B", "C")]
    reject_rows = [r for r in enriched if r["执行档位"] == "D"]
    b_rows = [r for r in enriched if r["执行档位"] == "B"]
    a_rows = [r for r in enriched if r["执行档位"] == "A"]
    p0_a = [r for r in a_rows if r["等级"] == "P0"]

    fields = [
        "用例ID", "用例标题", "等级", "用例类型", "模块", "用例目录", "所属模块",
        "执行档位", "档位说明", "技能分类", "匹配说明", "人工验收项",
        "前置条件", "步骤描述", "预期结果",
    ]

    write_csv(cases_dir / f"{stem}_可自动化用例.csv", fields, auto_rows)
    write_csv(cases_dir / f"{stem}_不可自动化用例.csv", fields, reject_rows)
    write_csv(cases_dir / f"{stem}_B档半自动.csv", fields, b_rows)

    share = detect_shared_steps(cases)
    write_csv(
        cases_dir / f"{stem}_共享步骤.csv",
        ["共享步骤名", "出现次数", "示例用例ID", "示例标题"],
        share,
    )

    mod_stats: dict[str, Counter[str]] = defaultdict(Counter)
    for r in enriched:
        mod_stats[r["模块"]][r["执行档位"]] += 1

    dep_lines = [
        f"# {stem} — 任务依赖与执行顺序",
        "",
        f"生成日期：{date.today().isoformat()}",
        f"源文件：`{input_path.relative_to(ROOT).as_posix()}`",
        f"规则：`{SKILL_MD.relative_to(ROOT).as_posix()}`",
        f"分类索引：`{MIDSCENE_INDEX.relative_to(ROOT).as_posix()}`",
        "",
        "## 概览",
        "",
        f"| 指标 | 数量 |",
        f"|------|------|",
        f"| 总用例 | {len(cases)} |",
        f"| A 档 | {tier_counts['A']} |",
        f"| B 档 | {tier_counts['B']} |",
        f"| C 档 | {tier_counts['C']} |",
        f"| D 档 | {tier_counts['D']} |",
        f"| P0 且 A 档 | {len(p0_a)} |",
        "",
        "## 产出目录",
        "",
        f"- 筛选结果：`{out_root.relative_to(ROOT).as_posix()}/`",
        f"- 终稿 YAML 建议：`{FINAL_YAML_DIR.relative_to(ROOT).as_posix()}/`",
        "",
    ]
    for k, v in sorted(skill_counts.items(), key=lambda x: -x[1]):
        dep_lines.append(f"- `{k}`：{v} 条")
    dep_lines.extend(["", "## 按模块分档（≥5 条）", ""])
    for mod, cnt in sorted(mod_stats.items(), key=lambda x: -sum(x[1].values())):
        if sum(cnt.values()) < 5:
            continue
        dep_lines.append(f"- **{mod}**：A={cnt['A']} B={cnt['B']} C={cnt['C']} D={cnt['D']}")

    report_dir.mkdir(parents=True, exist_ok=True)
    (report_dir / f"{stem}_依赖关系分析.md").write_text("\n".join(dep_lines) + "\n", encoding="utf-8")

    kw_stats: Counter[str] = Counter()
    for c in cases:
        for k in HIGH + MID:
            if k in c["步骤描述"]:
                kw_stats[k] += 1

    rpt = [
        f"# {stem} — 用例筛选报告",
        "",
        f"生成日期：{date.today().isoformat()}",
        f"源文件：`{input_path}`",
        "",
        "## 筛选结果",
        "",
        "| 口径 | 条数 | 占比 |",
        "|------|------|------|",
        f"| 总用例 | {len(cases)} | 100% |",
        f"| 可纳入自动化池（A+B+C） | {len(auto_rows)} | {100 * len(auto_rows) / max(len(cases), 1):.1f}% |",
        f"| D 档 | {len(reject_rows)} | {100 * len(reject_rows) / max(len(cases), 1):.1f}% |",
        "",
        "### 执行档位",
        "",
        "| 档位 | 条数 |",
        "|------|------|",
        f"| A | {tier_counts['A']} |",
        f"| B | {tier_counts['B']} |",
        f"| C | {tier_counts['C']} |",
        f"| D | {tier_counts['D']} |",
        "",
        "## 输出文件",
        "",
        f"- `{cases_dir.relative_to(ROOT)}/{stem}_可自动化用例.csv`",
        f"- `{cases_dir.relative_to(ROOT)}/{stem}_B档半自动.csv`",
        f"- `{cases_dir.relative_to(ROOT)}/{stem}_不可自动化用例.csv`",
        f"- `{drafts_dir.relative_to(ROOT)}/{stem}_自动化测试_草稿.yaml`",
        "",
        "## 下一步",
        "",
        "1. 评审 A 档；B 档见 `*_B档半自动.csv` 中「人工验收项」",
        f"2. 使用 **Midscene安卓自动化脚本** 将定稿写入 `{FINAL_YAML_DIR.relative_to(ROOT)}/`",
        "",
    ]
    for k, v in kw_stats.most_common(10):
        rpt.append(f"- 步骤关键词 `{k}`：{v} 条")

    (report_dir / f"{stem}_筛选报告.md").write_text("\n".join(rpt) + "\n", encoding="utf-8")

    drafts_dir.mkdir(parents=True, exist_ok=True)
    (drafts_dir / f"{stem}_自动化测试_草稿.yaml").write_text(
        build_yaml_draft(p0_a if p0_a else a_rows, limit=15, stem=stem),
        encoding="utf-8",
    )

    print(f"cases={len(cases)} A={tier_counts['A']} B={tier_counts['B']} C={tier_counts['C']} D={tier_counts['D']}")
    print(f"screen_out={out_root}")
    print(f"final_yaml_dir={FINAL_YAML_DIR}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Screen cases for Midscene (xlsx/csv)")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--out", type=Path, default=DEFAULT_SCREEN_OUT, help="筛选报告与 CSV 根目录")
    parser.add_argument("--stem", type=str, default=None)
    args = parser.parse_args()
    main(input_path=args.input, out_root=args.out, stem=args.stem)
